import yadisk
import openpyxl
from openpyxl import load_workbook

y = yadisk.YaDisk(token="y0_AgAAAABkLI6WAAhdVQAAAADNaD8OVl0R25QTRWmWQgEsSLm3RAF5ZSs")
y.download('/sverka.xlsx', "sverka.xlsx")  # cкачивание файла
wb = openpyxl.load_workbook(filename="sverka.xlsx")
wb.active = 0
sheetVlad = wb.active
wb.close()

def skach():
    y.download('/sverka.xlsx', "sverka.xlsx")  # cкачивание файла

def obnovl():
    wb = openpyxl.load_workbook(filename="sverka.xlsx")
    wb.active = 0
    sheetVlad = wb.active


def nom_counter():
    a2 = (sheetVlad['A2'].value)   #0
    a3 = (sheetVlad['A3'].value)   #1
    a4 = (sheetVlad['A4'].value)   #2
    a5 = (sheetVlad['A5'].value)   #3
    a6 = (sheetVlad['A6'].value)   #4
    a7 = (sheetVlad['A7'].value)   #5
    a8 = (sheetVlad['A8'].value)   #6
    a9 = (sheetVlad['A9'].value)   #7
    a10 = (sheetVlad['A10'].value) #8
    a11 = (sheetVlad['A11'].value) #9
    a12 = (sheetVlad['A12'].value) #10
    a13 = (sheetVlad['A13'].value) #11
    a14 = (sheetVlad['A14'].value) #12
    a15 = (sheetVlad['A15'].value) #13
    a16 = (sheetVlad['A16'].value) #14
    a17 = (sheetVlad['A17'].value) #15
    a18 = (sheetVlad['A18'].value) #16
    a19 = (sheetVlad['A19'].value) #17
    a20 = (sheetVlad['A20'].value) #18
    a21 = (sheetVlad['A21'].value)  # 18
    a22 = (sheetVlad['A22'].value)  # 18
    a23 = (sheetVlad['A23'].value)  # 18
    a24 = (sheetVlad['A24'].value)  # 18
    a25 = (sheetVlad['A25'].value)  # 18
    a26 = (sheetVlad['A26'].value)  # 18
    a27 = (sheetVlad['A27'].value)  # 18
    a28 = (sheetVlad['A28'].value)  # 18
    a29 = (sheetVlad['A29'].value)  # 18
    a30 = (sheetVlad['A30'].value)  # 18
    a31 = (sheetVlad['A31'].value)  # 18
    a32 = (sheetVlad['A32'].value)  # 18
    a33 = (sheetVlad['A33'].value)  # 18
    a34 = (sheetVlad['A34'].value)  # 18
    a35 = (sheetVlad['A35'].value)  # 18
    a36 = (sheetVlad['A36'].value)  # 18
    a37 = (sheetVlad['A37'].value)  # 18
    a38 = (sheetVlad['A38'].value)  # 18
    a39 = (sheetVlad['A39'].value)  # 18
    a40 = (sheetVlad['A40'].value)  # 18
    a41 = (sheetVlad['A41'].value)  # 18
    a42 = (sheetVlad['A42'].value)  # 18
    a43 = (sheetVlad['A43'].value)  # 18
    a44 = (sheetVlad['A44'].value)  # 18
    a45 = (sheetVlad['A45'].value)  # 18
    a46 = (sheetVlad['A46'].value)  # 18
    a47 = (sheetVlad['A47'].value)  # 18
    a48 = (sheetVlad['A48'].value)  # 18
    a49 = (sheetVlad['A49'].value)  # 18
    a50 = (sheetVlad['A50'].value)  # 18

    return [a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20,a21,a22,a23,a24,a25,a26,a27,a28,a29,a30,a31,a32,a33,a34,a35,a36,a37,a38,a39,a40,a41,a42,a43,a44,a45,a46,a47,a48,a49,a50]


def nom_name():
    wb = openpyxl.load_workbook(filename="sverka.xlsx")
    wb.active = 0
    sheetVlad = wb.active
    b2 = (sheetVlad['B2'].value)
    b3 = (sheetVlad['B3'].value)
    b4 = (sheetVlad['B4'].value)
    b5 = (sheetVlad['B5'].value)
    b6 = (sheetVlad['B6'].value)
    b7 = (sheetVlad['B7'].value)
    b8 = (sheetVlad['B8'].value)
    b9 = (sheetVlad['B9'].value)
    b10 = (sheetVlad['B10'].value)
    b11 = (sheetVlad['B11'].value)
    b12 = (sheetVlad['B12'].value)
    b13 = (sheetVlad['B13'].value)
    b14 = (sheetVlad['B14'].value)
    b15 = (sheetVlad['B15'].value)
    b16 = (sheetVlad['B16'].value)
    b17 = (sheetVlad['B17'].value)
    b18 = (sheetVlad['B18'].value)
    b19 = (sheetVlad['B19'].value)
    b20 = (sheetVlad['B20'].value)
    b21 = (sheetVlad['B21'].value)  # 18
    b22 = (sheetVlad['B22'].value)  # 18
    b23 = (sheetVlad['B23'].value)  # 18
    b24 = (sheetVlad['B24'].value)  # 18
    b25 = (sheetVlad['B25'].value)  # 18
    b26 = (sheetVlad['B26'].value)  # 18
    b27 = (sheetVlad['B27'].value)  # 18
    b28 = (sheetVlad['B28'].value)  # 18
    b29 = (sheetVlad['B29'].value)  # 18
    b30 = (sheetVlad['B30'].value)  # 18
    b31 = (sheetVlad['B31'].value)  # 18
    b32 = (sheetVlad['B32'].value)  # 18
    b33 = (sheetVlad['B33'].value)  # 18
    b34 = (sheetVlad['B34'].value)  # 18
    b35 = (sheetVlad['B35'].value)  # 18
    b36 = (sheetVlad['B36'].value)  # 18
    b37 = (sheetVlad['B37'].value)  # 18
    b38 = (sheetVlad['B38'].value)  # 18
    b39 = (sheetVlad['B39'].value)  # 18
    b40 = (sheetVlad['B40'].value)  # 18
    b41 = (sheetVlad['B41'].value)  # 18
    b42 = (sheetVlad['B42'].value)  # 18
    b43 = (sheetVlad['B43'].value)  # 18
    b44 = (sheetVlad['B44'].value)  # 18
    b45 = (sheetVlad['B45'].value)  # 18
    b46 = (sheetVlad['B46'].value)  # 18
    b47 = (sheetVlad['B47'].value)  # 18
    b48 = (sheetVlad['B48'].value)  # 18
    b49 = (sheetVlad['B49'].value)  # 18
    b50 = (sheetVlad['B50'].value)  # 18

    return [b2, b3, b4, b5, b6, b7, b8, b9, b10, b11, b12, b13, b14, b15, b16, b17, b18, b19, b20,b21,b22,b23,b24,b25,b26,b27,b28,b29,b30,b31,b32,b33,b34,b35,b36,b37,b38,b39,b40,b41,b42,b43,b44,b45,b46,b47,b48,b49,b50]







